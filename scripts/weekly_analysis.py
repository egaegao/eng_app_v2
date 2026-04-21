import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. SETUP DIREKTORI & KONFIGURASI
# ==========================================
BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

LOW_IS_BETTER = ["rain_hours", "slippery_hours", "fuel_ratio_mining", "stripping_ratio"]
PRIMARY_KPI_METRICS = ["overburden", "coal_getting", "rain_hours", "ewh_ob"]

# ==========================================
# 2. HELPER FUNCTIONS
# ==========================================
def parse_dates(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format="mixed", dayfirst=True, errors="coerce")
    return df


def safe_divide(numerator, denominator):
    if pd.isna(denominator) or denominator == 0:
        return 0
    return numerator / denominator


def compute_achievement(metric: str, actual: float, plan: float) -> float:
    return round(safe_divide(actual, plan) * 100, 1)


def traffic_light(metric: str, achievement: float) -> str:
    if metric in LOW_IS_BETTER:
        if achievement <= 100:
            return "GREEN"
        elif achievement <= 120:
            return "YELLOW"
        return "RED"
    else:
        if achievement >= 95:
            return "GREEN"
        elif achievement >= 80:
            return "YELLOW"
        return "RED"


def normalize_text_series(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.strip()
        .str.replace("_", " ", regex=False)
        .str.title()
    )


def get_status_priority(metric: str, achievement: float) -> float:
    if metric in LOW_IS_BETTER:
        return max(0, achievement - 100)
    return max(0, 100 - achievement)


def build_site_score(kpi_df: pd.DataFrame) -> float:
    if kpi_df.empty:
        return 0.0

    df = kpi_df.copy()

    def score_component(row):
        if row["metric"] in LOW_IS_BETTER:
            return max(0, min(120, 200 - row["achievement_pct"]))
        return max(0, min(120, row["achievement_pct"]))

    df["score_component"] = df.apply(score_component, axis=1)

    focus_df = df[df["metric"].isin(PRIMARY_KPI_METRICS)].copy()
    if focus_df.empty:
        focus_df = df.copy()

    return round(focus_df["score_component"].mean(), 1)


def score_label(score: float) -> str:
    if score >= 95:
        return "STRONG"
    elif score >= 80:
        return "WATCH"
    return "CRITICAL"


def auto_adjust_sheet_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_length = 0
        col_idx = col_cells[0].column
        col_letter = get_column_letter(col_idx)

        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = max(min_width, min(max_length + 2, max_width))
        ws.column_dimensions[col_letter].width = adjusted_width


def format_excel_header(ws, fill_color="1F4E78", font_color="FFFFFF"):
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def save_text_lines(filepath: Path, lines: list[str]):
    with open(filepath, "w", encoding="utf-8") as f:
        for line in lines:
            f.write(str(line).strip() + "\n")


def latest_non_null(series: pd.Series):
    s = series.dropna()
    if s.empty:
        return None
    return s.iloc[-1]


# ==========================================
# 3. MAIN ANALYSIS FUNCTION
# ==========================================
def run_analysis_from_data(
    weekly_kpi,
    fuel_ratio,
    issue_log,
    action_tracker,
    inventory,
    hauling_review,
    findings_summary,
):
    # --- 4. COERCION NUMERIC (STABILIZATION) ---
    for col in ["plan", "actual"]:
        if col in weekly_kpi.columns:
            weekly_kpi[col] = pd.to_numeric(weekly_kpi[col], errors="coerce")

    for col in ["actual_fr", "plan_fr", "actual_liter", "plan_liter"]:
        if col in fuel_ratio.columns:
            fuel_ratio[col] = pd.to_numeric(fuel_ratio[col], errors="coerce")

    for col in ["actual_ton", "target_ton", "achievement_ratio"]:
        if col in hauling_review.columns:
            hauling_review[col] = pd.to_numeric(hauling_review[col], errors="coerce")

    if "volume_ton" in inventory.columns:
        inventory["volume_ton"] = pd.to_numeric(inventory["volume_ton"], errors="coerce")

    # --- 5. DATA CLEANING & VALIDATION ---
    required_cols = ["metric", "plan", "actual", "week_date", "block"]
    missing_cols = [c for c in required_cols if c not in weekly_kpi.columns]
    if missing_cols:
        raise ValueError(f"Kolom wajib di weekly_kpi tidak lengkap: {missing_cols}")

    weekly_kpi = parse_dates(weekly_kpi, ["week_date"])
    fuel_ratio = parse_dates(fuel_ratio, ["week_date"])
    issue_log = parse_dates(issue_log, ["week_date"])
    action_tracker = parse_dates(action_tracker, ["week_date", "due_date"])
    inventory = parse_dates(inventory, ["week_date"])
    hauling_review = parse_dates(hauling_review, ["week_date"])

    weekly_kpi = weekly_kpi.dropna(subset=["week_date"]).copy()
    if weekly_kpi.empty:
        raise ValueError("weekly_kpi kosong atau seluruh week_date tidak valid.")

    if "severity" in issue_log.columns:
        issue_log["severity"] = issue_log["severity"].astype(str).str.strip().str.title()

    if "status" in issue_log.columns:
        issue_log["status"] = normalize_text_series(issue_log["status"])

    if "status" in action_tracker.columns:
        action_tracker["status"] = normalize_text_series(action_tracker["status"])

    if "priority" in action_tracker.columns:
        action_tracker["priority"] = action_tracker["priority"].astype(str).str.strip().str.title()

    if "inventory_type" in inventory.columns:
        inventory["inventory_type"] = normalize_text_series(inventory["inventory_type"])

    if "activity" in fuel_ratio.columns:
        fuel_ratio["activity"] = normalize_text_series(fuel_ratio["activity"])

    if "category" in findings_summary.columns:
        findings_summary["category"] = normalize_text_series(findings_summary["category"])

    # --- 6. KPI PROCESSING ---
    weekly_kpi["deviation"] = weekly_kpi["actual"] - weekly_kpi["plan"]
    weekly_kpi["gap"] = weekly_kpi["deviation"].round(2)
    
    weekly_kpi["achievement_pct"] = weekly_kpi.apply(
        lambda r: compute_achievement(r["metric"], r["actual"], r["plan"]), axis=1
    )
    weekly_kpi["status"] = weekly_kpi.apply(
        lambda r: traffic_light(r["metric"], r["achievement_pct"]), axis=1
    )
    weekly_kpi["priority_score"] = weekly_kpi.apply(
        lambda r: get_status_priority(r["metric"], r["achievement_pct"]), axis=1
    )

    kpi_summary = weekly_kpi.sort_values(["week_date", "block", "metric"]).copy()
    kpi_summary.to_csv(OUTPUT_DIR / "kpi_summary.csv", index=False)

    kpi_block_summary = (
        weekly_kpi.groupby("block", as_index=False)
        .agg(
            total_plan=("plan", "sum"),
            total_actual=("actual", "sum"),
            avg_achievement_pct=("achievement_pct", "mean"),
            avg_priority_score=("priority_score", "mean"),
            latest_week=("week_date", "max"),
        )
    )
    
    site_health_data = (
        weekly_kpi.groupby("block")
        .apply(build_site_score)
        .reset_index(name="site_health_score")
    )
    kpi_block_summary = kpi_block_summary.merge(site_health_data, on="block")
    
    kpi_block_summary["site_health_label"] = kpi_block_summary["site_health_score"].apply(score_label)
    kpi_block_summary["avg_achievement_pct"] = kpi_block_summary["avg_achievement_pct"].round(1)
    kpi_block_summary["avg_priority_score"] = kpi_block_summary["avg_priority_score"].round(1)
    kpi_block_summary.to_csv(OUTPUT_DIR / "kpi_block_summary.csv", index=False)

    kpi_priority_summary = (
        weekly_kpi.sort_values(["week_date", "block", "priority_score"], ascending=[False, True, False])
        .copy()
    )
    kpi_priority_summary.to_csv(OUTPUT_DIR / "kpi_priority_summary.csv", index=False)

    # --- 7. ISSUE LOG PROCESSING ---
    high_issues = issue_log[issue_log["severity"] == "High"].copy()
    high_issues.to_csv(OUTPUT_DIR / "high_issues.csv", index=False)

    issue_status_summary = (
        issue_log.groupby(["block", "status"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
    )
    issue_status_summary.to_csv(OUTPUT_DIR / "issue_status_summary.csv", index=False)

    issue_category_summary = (
        issue_log.groupby(["block", "issue_category", "severity"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
        .sort_values(["block", "count"], ascending=[True, False])
    )
    issue_category_summary.to_csv(OUTPUT_DIR / "issue_category_summary.csv", index=False)

    issue_impact_summary = (
        issue_log.groupby(["block", "impact_area"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
        .sort_values(["block", "count"], ascending=[True, False])
    )
    issue_impact_summary.to_csv(OUTPUT_DIR / "issue_impact_summary.csv", index=False)

    # --- 8. ACTION TRACKER PROCESSING ---
    action_tracker["days_to_due"] = (
        action_tracker["due_date"] - action_tracker["week_date"]
    ).dt.days

    action_tracker["overdue_flag"] = (
        (action_tracker["due_date"].notna()) &
        (action_tracker["week_date"].notna()) &
        (action_tracker["status"].isin(["Open", "Progress"])) &
        (action_tracker["due_date"] < action_tracker["week_date"])
    )

    action_status_summary = (
        action_tracker.groupby(["block", "status"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
    )
    action_status_summary.to_csv(OUTPUT_DIR / "action_status_summary.csv", index=False)

    action_priority_summary = (
        action_tracker.groupby(["block", "priority"], as_index=False)
        .size()
        .rename(columns={"size": "count"})
    )
    action_priority_summary.to_csv(OUTPUT_DIR / "action_priority_summary.csv", index=False)

    overdue_actions = action_tracker[action_tracker["overdue_flag"]].copy()
    overdue_actions.to_csv(OUTPUT_DIR / "overdue_actions.csv", index=False)

    # --- 9. INVENTORY PROCESSING ---
    inventory_summary = (
        inventory.groupby(["block", "inventory_type"], as_index=False)["volume_ton"]
        .sum()
        .sort_values(["block", "inventory_type"])
    )
    inventory_summary.to_csv(OUTPUT_DIR / "inventory_summary.csv", index=False)

    inventory_seam_summary = (
        inventory.groupby(["block", "seam", "inventory_type"], as_index=False)["volume_ton"]
        .sum()
        .sort_values(["block", "seam", "inventory_type"])
    )
    inventory_seam_summary.to_csv(OUTPUT_DIR / "inventory_seam_summary.csv", index=False)

    inventory_status_summary = (
        inventory.groupby(["block", "status_note"], as_index=False)["volume_ton"]
        .sum()
        .sort_values(["block", "volume_ton"], ascending=[True, False])
    )
    inventory_status_summary.to_csv(OUTPUT_DIR / "inventory_status_summary.csv", index=False)

    # --- 10. FUEL RATIO PROCESSING ---
    fuel_ratio["fr_dev"] = fuel_ratio["actual_fr"] - fuel_ratio["plan_fr"]
    fuel_ratio["liter_dev"] = fuel_ratio["actual_liter"] - fuel_ratio["plan_liter"]
    fuel_ratio["fr_achievement_pct"] = fuel_ratio.apply(
        lambda r: compute_achievement("fuel_ratio_mining", r["actual_fr"], r["plan_fr"]), axis=1
    )
    fuel_ratio["fuel_status"] = fuel_ratio["fr_achievement_pct"].apply(
        lambda x: traffic_light("fuel_ratio_mining", x)
    )
    fuel_ratio.to_csv(OUTPUT_DIR / "fuel_ratio_summary.csv", index=False)

    fuel_ratio_block_summary = (
        fuel_ratio.groupby(["block", "activity"], as_index=False)
        .agg(
            plan_liter=("plan_liter", "sum"),
            actual_liter=("actual_liter", "sum"),
            plan_fr=("plan_fr", "mean"),
            actual_fr=("actual_fr", "mean"),
            avg_fr_achievement_pct=("fr_achievement_pct", "mean"),
        )
    )
    fuel_ratio_block_summary["liter_dev"] = fuel_ratio_block_summary["actual_liter"] - fuel_ratio_block_summary["plan_liter"]
    fuel_ratio_block_summary["fr_dev"] = fuel_ratio_block_summary["actual_fr"] - fuel_ratio_block_summary["plan_fr"]
    fuel_ratio_block_summary["fuel_status"] = fuel_ratio_block_summary["avg_fr_achievement_pct"].apply(
        lambda x: traffic_light("fuel_ratio_mining", x)
    )
    fuel_ratio_block_summary["avg_fr_achievement_pct"] = fuel_ratio_block_summary["avg_fr_achievement_pct"].round(1)
    fuel_ratio_block_summary["plan_fr"] = fuel_ratio_block_summary["plan_fr"].round(2)
    fuel_ratio_block_summary["actual_fr"] = fuel_ratio_block_summary["actual_fr"].round(2)
    fuel_ratio_block_summary["fr_dev"] = fuel_ratio_block_summary["fr_dev"].round(2)
    fuel_ratio_block_summary.to_csv(OUTPUT_DIR / "fuel_ratio_block_summary.csv", index=False)

    # --- 11. HAULING REVIEW PROCESSING ---
    hauling_review["achievement_pct"] = pd.to_numeric(
        hauling_review["achievement_ratio"], errors="coerce"
    )
    hauling_review["achievement_pct"] = hauling_review["achievement_pct"].apply(
        lambda x: x * 100 if pd.notna(x) and x <= 1.5 else x
    )

    mask_missing_ratio = hauling_review["achievement_pct"].isna()
    hauling_review.loc[mask_missing_ratio, "achievement_pct"] = (
        hauling_review.loc[mask_missing_ratio]
        .apply(lambda r: round(safe_divide(r["actual_ton"], r["target_ton"]) * 100, 1), axis=1)
    )

    hauling_review["ton_dev"] = hauling_review["actual_ton"] - hauling_review["target_ton"]
    hauling_review["status"] = hauling_review["achievement_pct"].apply(
        lambda x: "GREEN" if x >= 95 else ("YELLOW" if x >= 80 else "RED")
    )
    hauling_review.to_csv(OUTPUT_DIR / "hauling_review_summary.csv", index=False)

    hauling_route_summary = (
        hauling_review.groupby("route", as_index=False)
        .agg(
            target_ton=("target_ton", "sum"),
            actual_ton=("actual_ton", "sum"),
            avg_achievement_pct=("achievement_pct", "mean"),
        )
    )
    hauling_route_summary["ton_dev"] = hauling_route_summary["actual_ton"] - hauling_route_summary["target_ton"]
    hauling_route_summary["status"] = hauling_route_summary["avg_achievement_pct"].apply(
        lambda x: "GREEN" if x >= 95 else ("YELLOW" if x >= 80 else "RED")
    )
    hauling_route_summary["avg_achievement_pct"] = hauling_route_summary["avg_achievement_pct"].round(1)
    hauling_route_summary.to_csv(OUTPUT_DIR / "hauling_route_summary.csv", index=False)

    # --- 12. FINDINGS PROCESSING (UPDATED: DETAIL-BASED) ---
    findings_summary_proc = findings_summary.copy()

    # Normalisasi status
    if "status" in findings_summary_proc.columns:
        findings_summary_proc["status"] = normalize_text_series(findings_summary_proc["status"])

    # Hitung summary dari detail (Count per Status)
    if "status" in findings_summary_proc.columns:
        status_counts = findings_summary_proc["status"].value_counts()

        total_open = int(status_counts.get("Open", 0))
        total_progress = int(status_counts.get("Progress", 0))
        total_closed = int(status_counts.get("Closed", 0))
        total_findings = int(len(findings_summary_proc))
    else:
        total_open = total_progress = total_closed = total_findings = 0

    # Simpan processed (copy detail)
    findings_summary_proc.to_csv(OUTPUT_DIR / "findings_summary_processed.csv", index=False)

    # Summary overall baru yang dihitung otomatis
    findings_overall_summary = pd.DataFrame([{
        "total_open": total_open,
        "total_closed": total_closed,
        "total_progress": total_progress,
        "total_findings": total_findings
    }])

    findings_overall_summary.to_csv(OUTPUT_DIR / "findings_overall_summary.csv", index=False)

    # --- 13. VISUALIZATION (CHARTS) ---
    metrics_to_plot = [
        ("overburden", "BCM", "Overburden Plan vs Actual", "chart_overburden.png"),
        ("coal_getting", "Ton", "Coal Getting Plan vs Actual", "chart_coal_getting.png"),
        ("rain_hours", "Hours", "Rain Hours Plan vs Actual", "chart_rain_hours.png"),
        ("ewh_ob", "Hours", "EWH OB Plan vs Actual", "chart_ewh_ob.png")
    ]

    for metric, unit, title, filename in metrics_to_plot:
        data = weekly_kpi[weekly_kpi["metric"] == metric].copy()
        if not data.empty:
            data = data.sort_values(["week_date", "block"]).reset_index(drop=True)
            data["label"] = data["block"].astype(str)
            if "week_date" in data.columns and data["week_date"].notna().any():
                data["label"] = data.apply(
                    lambda r: f"{r['block']} | {r['week_date'].strftime('%d-%b')}" if pd.notna(r["week_date"]) else str(r["block"]),
                    axis=1
                )

            plt.figure(figsize=(10, 5))
            x = range(len(data))
            plt.bar(x, data["plan"], width=0.4, label="Plan")
            plt.bar([i + 0.4 for i in x], data["actual"], width=0.4, label="Actual")
            plt.xticks([i + 0.2 for i in x], data["label"], rotation=45, ha="right")
            plt.ylabel(unit)
            plt.title(title)
            plt.legend()
            plt.tight_layout()
            plt.savefig(OUTPUT_DIR / filename, dpi=150)
            plt.close()

    if not hauling_route_summary.empty:
        plt.figure(figsize=(10, 5))
        x = range(len(hauling_route_summary))
        plt.bar(x, hauling_route_summary["target_ton"], width=0.4, label="Target")
        plt.bar([i + 0.4 for i in x], hauling_route_summary["actual_ton"], width=0.4, label="Actual")
        plt.xticks([i + 0.2 for i in x], hauling_route_summary["route"], rotation=30, ha="right")
        plt.ylabel("Ton")
        plt.title("Hauling Route Target vs Actual")
        plt.legend()
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / "chart_hauling_review.png", dpi=150)
        plt.close()

    if not inventory_summary.empty:
        pivot_inventory = inventory_summary.pivot(index="block", columns="inventory_type", values="volume_ton").fillna(0)
        pivot_inventory.plot(kind="bar", figsize=(10, 5))
        plt.ylabel("Volume (Ton)")
        plt.title("Inventory by Block and Type")
        plt.xticks(rotation=30, ha="right")
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / "chart_inventory_summary.png", dpi=150)
        plt.close()

    # --- 14. EXECUTIVE SUMMARY GENERATOR ---
    def generate_executive_summary_internal():
        lines = []
        if weekly_kpi.empty:
            return ["No weekly KPI data available."]

        latest_week = latest_non_null(weekly_kpi["week_date"])
        if latest_week is not None:
            lines.append(f"Weekly report based on latest data up to {latest_week.strftime('%d %b %Y')}.")

        for block in sorted(weekly_kpi["block"].dropna().astype(str).unique()):
            sub = weekly_kpi[weekly_kpi["block"].astype(str) == block].copy()
            if sub.empty: continue
            latest_block_week = latest_non_null(sub["week_date"])
            if latest_block_week is not None:
                sub = sub[sub["week_date"] == latest_block_week].copy()

            def pick_metric(metric_name):
                temp = sub[sub["metric"] == metric_name]
                return temp.iloc[0] if not temp.empty else None

            ob = pick_metric("overburden")
            coal = pick_metric("coal_getting")
            rain = pick_metric("rain_hours")
            ewh = pick_metric("ewh_ob")

            sentence = f"Block {block}: "
            if ob is not None:
                if ob["achievement_pct"] < 80: sentence += f"overburden underperformed at {ob['achievement_pct']:.1f}% of plan, "
                elif ob["achievement_pct"] < 95: sentence += f"overburden was near target at {ob['achievement_pct']:.1f}% of plan, "
                else: sentence += f"overburden performed well at {ob['achievement_pct']:.1f}% of plan, "
            if coal is not None:
                if coal["achievement_pct"] < 80: sentence += f"coal getting was below target at {coal['achievement_pct']:.1f}%, "
                else: sentence += f"coal getting remained relatively stable at {coal['achievement_pct']:.1f}%, "
            if rain is not None and rain["actual"] > rain["plan"]:
                sentence += f"with rain hours above plan ({rain['actual']:.1f} vs {rain['plan']:.1f}), "
            if ewh is not None and ewh["achievement_pct"] < 80:
                sentence += f"and EWH OB was low at {ewh['achievement_pct']:.1f}%. "
            else:
                sentence = sentence.rstrip(", ") + "."
            lines.append(sentence)

        if not high_issues.empty:
            top_issue = high_issues.groupby("issue_category").size().sort_values(ascending=False)
            lines.append(f"High severity issue paling dominan adalah '{top_issue.index[0]}' sebanyak {top_issue.values[0]} kasus.")

        if not action_status_summary.empty:
            acts = action_status_summary.set_index("status")["count"].to_dict()
            lines.append(f"Action tracker menunjukkan Open={int(acts.get('Open',0))}, Progress={int(acts.get('Progress',0))}, Closed={int(acts.get('Closed',0))}.")

        if not fuel_ratio_block_summary.empty:
            worst_fuel = fuel_ratio_block_summary.sort_values("avg_fr_achievement_pct", ascending=False).iloc[0]
            lines.append(f"Fuel ratio tertinggi terdapat pada block {worst_fuel['block']} aktivitas {worst_fuel['activity']} dengan pencapaian {worst_fuel['avg_fr_achievement_pct']:.1f}% terhadap plan.")

        if not inventory_summary.empty:
            top_inv = inventory_summary.sort_values("volume_ton", ascending=False).iloc[0]
            lines.append(f"Inventory terbesar berada di block {top_inv['block']} untuk tipe {top_inv['inventory_type']} sebesar {top_inv['volume_ton']:,.0f} ton.")

        if not hauling_route_summary.empty:
            worst_r = hauling_route_summary.sort_values("avg_achievement_pct").iloc[0]
            lines.append(f"Rute hauling dengan performa terendah adalah {worst_r['route']} pada {worst_r['avg_achievement_pct']:.1f}% dengan deviasi {worst_r['ton_dev']:,.0f} ton.")

        if not findings_overall_summary.empty:
            fr = findings_overall_summary.iloc[0]
            lines.append(f"Findings summary mencatat total {int(fr['total_findings'])} temuan: Open={int(fr['total_open'])}, Progress={int(fr['total_progress'])}, Closed={int(fr['total_closed'])}.")

        return lines

    executive_summary_lines = generate_executive_summary_internal()
    save_text_lines(OUTPUT_DIR / "executive_summary.txt", executive_summary_lines)
    summary_df = pd.DataFrame({"Executive Summary": executive_summary_lines})

    # --- 15. EXPORT TO EXCEL REPORT ---
    excel_path = OUTPUT_DIR / "weekly_report.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        kpi_summary.to_excel(writer, sheet_name="KPI Summary", index=False)
        kpi_block_summary.to_excel(writer, sheet_name="KPI Block Summary", index=False)
        kpi_priority_summary.to_excel(writer, sheet_name="KPI Priorities", index=False)
        high_issues.to_excel(writer, sheet_name="High Issues", index=False)
        issue_status_summary.to_excel(writer, sheet_name="Issue Status Summary", index=False)
        issue_category_summary.to_excel(writer, sheet_name="Issue Category Summary", index=False)
        issue_impact_summary.to_excel(writer, sheet_name="Issue Impact Summary", index=False)
        action_tracker.to_excel(writer, sheet_name="Action Tracker", index=False)
        action_status_summary.to_excel(writer, sheet_name="Action Status Summary", index=False)
        action_priority_summary.to_excel(writer, sheet_name="Action Priority Summary", index=False)
        overdue_actions.to_excel(writer, sheet_name="Overdue Actions", index=False)
        inventory_summary.to_excel(writer, sheet_name="Inventory Summary", index=False)
        inventory_seam_summary.to_excel(writer, sheet_name="Inventory Seam Summary", index=False)
        inventory_status_summary.to_excel(writer, sheet_name="Inventory Status Summary", index=False)
        fuel_ratio.to_excel(writer, sheet_name="Fuel Analysis", index=False)
        fuel_ratio_block_summary.to_excel(writer, sheet_name="Fuel Block Summary", index=False)
        hauling_review.to_excel(writer, sheet_name="Hauling Review", index=False)
        hauling_route_summary.to_excel(writer, sheet_name="Hauling Route Summary", index=False)
        findings_summary_proc.to_excel(writer, sheet_name="Findings Summary", index=False)
        findings_overall_summary.to_excel(writer, sheet_name="Findings Overall", index=False)

    # --- 16. FORMAT EXCEL (STYLING) ---
    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        format_excel_header(ws)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        auto_adjust_sheet_width(ws, min_width=10, max_width=40)

    ws_exec = wb["Executive Summary"]
    ws_exec["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_exec.column_dimensions["A"].width = 120
    wb.save(excel_path)

    # --- 17. RETURN DATA ---
    return {
        "kpi_summary": kpi_summary,
        "kpi_block_summary": kpi_block_summary,
        "fuel_ratio": fuel_ratio,
        "hauling_review": hauling_review,
        "inventory_summary": inventory_summary,
        "executive_summary": executive_summary_lines,
    }